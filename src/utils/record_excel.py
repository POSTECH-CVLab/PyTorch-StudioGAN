import openpyxl
import argparse
import os

# log must be in order of train-legacy
def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("--log_path", type=str, help="directory that includes checkpoints", required=True)
    parser.add_argument("--dataset", type=str, help="dataset in [CIFAR10, Tiny_ImageNet, ImageNet, AFHQ, CUB200", required=True)
    parser.add_argument("--num_eval", type=int, help="how many time to repeat", required=True)
    args = parser.parse_args()
    return args

def fit_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def write_excel(parsed_args):
    log_list = os.listdir(parsed_args.log_path)
    num_eval = parsed_args.num_eval
    wb = openpyxl.Workbook()
    wb.create_sheet(parsed_args.dataset, 0)
    wb.active = 0
    sheet = wb.active
    metrics = ["IS", "FID", "Improved Precision", "Improved Recall", "Density", "Coverage"]
    row, col = 2, 1
    for metric in metrics:
        sheet.cell(row,col).value = metric + "_avg"
        row += 1
    for metric in metrics:
        sheet.cell(row,col).value = metric + "_std"
        row += 1
    row += 2
    for metric in metrics:
        for i in range(1, num_eval+1):
            sheet.cell(row,col).value = metric + "_" + str(i)
            row += 1
    
    valid = "valid" if parsed_args.dataset == "ImageNet" else "test"
    for log in log_list:
        with open(os.path.join(parsed_args.log_path, log), 'r') as f:
            lines = f.readlines()
            num_run = 0
            for line in lines:
                line = line.strip()
                if "> ------------------------------------" in line:
                    num_run += 1

                if "Start Evaluation ("in line:
                    if num_run % num_eval == 1:
                        col += 1
                        is_row, fid_row, precision_row, recall_row, density_row, coverage_row = 16, 16 + 1 * num_eval, 16 + 2 * num_eval, 16 + 3 * num_eval, 16 + 4 * num_eval, 16 + 5 * num_eval
                        if (num_run - 1) / num_eval == 0:
                            sheet.cell(1, col).value = line[line.rindex(": ")+1:line.rindex("-train-")] + "-"+valid+"-legacy"
                        if (num_run - 1) / num_eval == 1:
                            sheet.cell(1, col).value = line[line.rindex(": ")+1:line.rindex("-train-")] + "-"+valid+"-clean"
                        if (num_run - 1) / num_eval == 2:
                            sheet.cell(1, col).value = line[line.rindex(": ")+1:line.rindex("-train-")] + "-train-legacy"
                        if (num_run - 1) / num_eval == 3:
                            sheet.cell(1, col).value = line[line.rindex(": ")+1:line.rindex("-train-")] + "-train-clean"
                        sheet.cell(2, col).value = "=AVERAGE(" + sheet.cell(2, col).column_letter + str(is_row) + ":" + sheet.cell(2, col).column_letter + str(is_row + num_eval - 1) + ")"
                        sheet.cell(3, col).value = "=AVERAGE(" + sheet.cell(2, col).column_letter + str(fid_row) + ":" + sheet.cell(2, col).column_letter + str(fid_row + num_eval - 1) + ")"
                        sheet.cell(4, col).value = "=AVERAGE(" + sheet.cell(2, col).column_letter + str(precision_row) + ":" + sheet.cell(2, col).column_letter + str(precision_row + num_eval - 1) + ")"
                        sheet.cell(5, col).value = "=AVERAGE(" + sheet.cell(2, col).column_letter + str(recall_row) + ":" + sheet.cell(2, col).column_letter + str(recall_row + num_eval - 1) + ")"
                        sheet.cell(6, col).value = "=AVERAGE(" + sheet.cell(2, col).column_letter + str(density_row) + ":" + sheet.cell(2, col).column_letter + str(density_row + num_eval - 1) + ")"
                        sheet.cell(7, col).value = "=AVERAGE(" + sheet.cell(2, col).column_letter + str(coverage_row) + ":" + sheet.cell(2, col).column_letter + str(coverage_row + num_eval - 1) + ")"
                        sheet.cell(8, col).value = "=STDEV(" + sheet.cell(2, col).column_letter + str(is_row) + ":" + sheet.cell(2, col).column_letter + str(is_row + num_eval - 1) + ")"
                        sheet.cell(9, col).value = "=STDEV(" + sheet.cell(2, col).column_letter + str(fid_row) + ":" + sheet.cell(2, col).column_letter + str(fid_row + num_eval - 1) + ")"
                        sheet.cell(10, col).value = "=STDEV(" + sheet.cell(2, col).column_letter + str(precision_row) + ":" + sheet.cell(2, col).column_letter + str(precision_row + num_eval - 1) + ")"
                        sheet.cell(11, col).value = "=STDEV(" + sheet.cell(2, col).column_letter + str(recall_row) + ":" + sheet.cell(2, col).column_letter + str(recall_row + num_eval - 1) + ")"
                        sheet.cell(12, col).value = "=STDEV(" + sheet.cell(2, col).column_letter + str(density_row) + ":" + sheet.cell(2, col).column_letter + str(density_row + num_eval - 1) + ")"
                        sheet.cell(13, col).value = "=STDEV(" + sheet.cell(2, col).column_letter + str(coverage_row) + ":" + sheet.cell(2, col).column_letter + str(coverage_row + num_eval - 1) + ")"
                
                elif "Inception score (" in line:
                    sheet.cell(is_row, col).value = float(line[line.rindex(": ")+1:])
                    is_row += 1
                elif "FID score (" in line:
                    sheet.cell(fid_row, col).value = float(line[line.rindex(": ")+1:])
                    fid_row += 1
                elif "Improved Precision (" in line:
                    sheet.cell(precision_row, col).value = float(line[line.rindex(": ")+1:])
                    precision_row += 1
                elif "Improved Recall (" in line:
                    sheet.cell(recall_row, col).value = float(line[line.rindex(": ")+1:])
                    recall_row += 1
                elif "Density (" in line:
                    sheet.cell(density_row, col).value = float(line[line.rindex(": ")+1:])
                    density_row += 1
                elif "Coverage (" in line:
                    sheet.cell(coverage_row, col).value = float(line[line.rindex(": ")+1:])
                    coverage_row += 1
    fit_width(sheet)
    wb.save(parsed_args.dataset+"_eval_results.xlsx")

if __name__ == "__main__":
    parsed_args = parse_arguments()
    write_excel(parsed_args)